from pathlib import Path
import os
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import pandas as pd
from openpyxl.utils import get_column_letter


APP_FOLDER = "ProcesadorCSV"

def base_output_dir() -> Path:
    docs = Path(os.path.expanduser("~")) / "Documents"
    return docs / APP_FOLDER / "OUTPUT"

def make_output_path(test_name: str, input_file: Path, ext: str = ".xlsx") -> Path:
    out_dir = base_output_dir() / test_name
    out_dir.mkdir(parents=True, exist_ok=True)

    safe_test = test_name.replace(" ", "_")
    base = out_dir / f"{input_file.stem}__{safe_test}{ext}"

    if not base.exists():
        return base

    i = 1
    while True:
        cand = out_dir / f"{input_file.stem}__{safe_test}_{i}{ext}"
        if not cand.exists():
            return cand
        i += 1


def reveal_in_explorer(path: Path) -> None:
    if path.is_dir():
        os.startfile(str(path))
    else:
        subprocess.run(["explorer", "/select,", str(path)], check=False)

def aplicar_bordes_excel(path: Path):
    wb = load_workbook(path)
    ws = wb.active

    thin = Side(style="thin")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = borde

    wb.save(path)
def auto_ajustar_columnas(path: Path, padding: int = 2):
    wb = load_workbook(path)
    ws = wb.active

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value is not None:
                val = str(cell.value)
                max_len = max(max_len, len(val))

        ws.column_dimensions[col_letter].width = max_len + padding

    wb.save(path)