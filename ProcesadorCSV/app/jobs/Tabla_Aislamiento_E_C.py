from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.cell.cell import MergedCell
from app.output import base_output_dir
from app.output import aplicar_bordes_excel
from app.output import auto_ajustar_columnas

TEST_NAME = "Tabla_Aislamiento_E_C"


def nombre_salida_unico(path_out: Path) -> Path:
    if not path_out.exists():
        return path_out
    i = 1
    while True:
        cand = path_out.with_stem(f"{path_out.stem}_{i}")
        if not cand.exists():
            return cand
        i += 1


def mismo_diferencial(c1: int, c2: int, diferenciales: dict[int, list[int]]) -> bool:
    for circuitos in diferenciales.values():
        if c1 in circuitos and c2 in circuitos:
            return True
    return False


def run(circuitos_por_dif: list[int]) -> Path:


    diferenciales = {}
    c_global = 1
    for d, n_c in enumerate(circuitos_por_dif, start=1):
        n_c = int(n_c)
        diferenciales[d] = list(range(c_global, c_global + n_c))
        c_global += n_c
    C_cir = c_global - 1 

    cols = ["Circuito_A", "FNT_A", "Circuito_B", "FNT_B", "Resistencia", "SI", "NO", "Observación"]
    df = pd.DataFrame(columns=cols)

    def add_row(**kwargs):
        nonlocal df
        df.loc[len(df)] = {c: kwargs.get(c, "") for c in cols}

    for i in range(1, C_cir + 1):
        for k in range(i + 1, C_cir + 1):
            add_row(Circuito_A=str(i), FNT_A="F", Circuito_B=str(k), FNT_B="F", Resistencia="")
            add_row(Circuito_A=str(i), FNT_A="F", Circuito_B=str(k), FNT_B="N", Resistencia="")
            add_row(Circuito_A=str(i), FNT_A="N", Circuito_B=str(k), FNT_B="F", Resistencia="")

            if mismo_diferencial(i, k, diferenciales):
                add_row(
                    Circuito_A=str(i), FNT_A="N", Circuito_B=str(k), FNT_B="N",
                    Resistencia="N/A", SI="N/A", NO="N/A", Observación="Mismo diferencial"
                )
            else:
                add_row(Circuito_A=str(i), FNT_A="N", Circuito_B=str(k), FNT_B="N", Resistencia="")

    for i in range(1, C_cir + 1):
        add_row(Circuito_A="Barra", FNT_A="T", Circuito_B=str(i), FNT_B="F", Resistencia="")

        if i > 1 and mismo_diferencial(i, i - 1, diferenciales):
            add_row(
                Circuito_A="Barra", FNT_A="T", Circuito_B=str(i), FNT_B="N",
                Resistencia="N/A", SI="N/A", NO="N/A", Observación="Mismo diferencial"
            )
        else:
            add_row(Circuito_A="Barra", FNT_A="T", Circuito_B=str(i), FNT_B="N", Resistencia="")

    top = [
        "Circuito A", "Circuito A",
        "Circuito B", "Circuito B",
        "Medición",
        "CONFORME", "CONFORME", "CONFORME"
    ]
    bottom = [
        "Circuito", "F/N/T",
        "Circuito", "F/N/T",
        "Resistencia",
        "SI", "NO", "Observación"
    ]
    df.columns = pd.MultiIndex.from_arrays([top, bottom])

    out_dir = base_output_dir() / TEST_NAME
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = nombre_salida_unico(out_dir / "Tabla_A_E_C.xlsx")

    df.index.name = ""
    df.to_excel(out_path, index=True)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    ws["A1"].value = None
    ws["A2"].value = None
    ws.column_dimensions["A"].hidden = True

    start_col = 2
    end_col = start_col + df.shape[1] - 1

    for c in range(start_col, end_col + 1):
        sub_cell = ws.cell(row=2, column=c)
        if isinstance(sub_cell, MergedCell):
            continue
        sub = sub_cell.value
        if sub is None or str(sub).strip() == "":
            sub_cell.value = None
            ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

    row_extra = df.columns.nlevels + 1 
    ws.delete_rows(row_extra)

    wb.save(out_path)
    auto_ajustar_columnas(out_path)
    aplicar_bordes_excel(out_path)
    return out_path
