from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from itertools import combinations
from app.output import make_output_path
from app.output import aplicar_bordes_excel
from app.output import auto_ajustar_columnas


TEST_NAME = "Tabla_Riso_Entre_Circuitos"

CONDUCTORES = ["R", "S", "T", "N", "Ti"]
COMBOS_MISMO_DIF = [("R", "S"), ("R", "T"), ("S", "T")]
COMBOS_DISTINTO_DIF = list(combinations(CONDUCTORES, 2))

CONFORME = ["SI", "NO", "N/A", "Observación"]


def _build_circuito_a_dif(nc: list[int]) -> dict[int, int]:

    circuito_a_dif = {}
    circuito_global = 1

    for dif_idx, n in enumerate(nc, start=1):
        n = int(n)
        if n < 0:
            raise ValueError("Circuitos por diferencial no puede ser negativo.")
        for _ in range(n):
            circuito_a_dif[circuito_global] = dif_idx
            circuito_global += 1

    return circuito_a_dif


def run(nc: list[int]) -> Path:

    if not nc:
        raise ValueError("Debes ingresar al menos 1 diferencial.")
    if any(int(x) < 0 for x in nc):
        raise ValueError("Circuitos por diferencial inválidos (enteros >= 0).")

    circuito_a_dif = _build_circuito_a_dif(nc)
    total_cir = max(circuito_a_dif.keys(), default=0)

    if total_cir < 2:
        raise ValueError("Necesitas al menos 2 circuitos totales para generar combinaciones.")

    filas = []
    for a in range(1, total_cir + 1):
        for b in range(a + 1, total_cir + 1):
            dif_a = circuito_a_dif.get(a)
            dif_b = circuito_a_dif.get(b)

            mismos = (dif_a == dif_b)
            combos = COMBOS_MISMO_DIF if mismos else COMBOS_DISTINTO_DIF

            for c1, c2 in combos:
                row = {
                    "Circuito A": a,
                    "Dif A": dif_a,
                    "Circuito B": b,
                    "Dif B": dif_b,
                    "Combinación": f"{c1}-{c2}",
                }
                for cond in CONDUCTORES:
                    row[cond] = "X" if cond in (c1, c2) else ""
                filas.append(row)

    df = pd.DataFrame(
        filas,
        columns=["Circuito A", "Dif A", "Circuito B", "Dif B", "Combinación"] + CONDUCTORES,
    )

    out_path = make_output_path(TEST_NAME, Path("TABLA.xlsx"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    headers = list(df.columns)

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c).value = h
        ws.cell(row=2, column=c).value = ""

    start = len(headers) + 1
    for i, sub in enumerate(CONFORME):
        ws.cell(row=1, column=start + i).value = "CONFORME"
        ws.cell(row=2, column=start + i).value = sub

    ws.merge_cells(
        start_row=1,
        start_column=start,
        end_row=1,
        end_column=start + len(CONFORME) - 1
    )

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for r in (1, 2):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = bold
            cell.alignment = center

    ws.freeze_panes = "A3"

    for r, row in enumerate(df.itertuples(index=False), start=3):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c)
            cell.value = v
            if v == "X":
                cell.alignment = center

    wb.save(out_path)
    aplicar_bordes_excel(out_path)
    auto_ajustar_columnas(out_path)
    return out_path
