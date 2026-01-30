from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.cell.cell import MergedCell
from app.output import make_output_path
from app.output import base_output_dir
from app.output import aplicar_bordes_excel
from app.output import auto_ajustar_columnas

TEST_NAME = "Tabla_Continuidad"


def add_row(df, row_dict):
    df.loc[len(df)] = row_dict
    return df


def run(al: int, cargas_por_circuito: list[int]) -> Path:

    cir = len(cargas_por_circuito)

    df = pd.DataFrame(columns=[
        "Circuito", "Carga", "Configuración",
        "Resistencia (+)", "Resistencia (-)",
        "SI", "NO", "Observación"
    ])

    for i in range(cir):
        cargas = int(cargas_por_circuito[i])

        for k in range(cargas):
            df = add_row(df, {"Circuito": f"N°{i+1}", "Carga": f"{k+1}", "Configuración": "N-PE",
                              "Resistencia (+)": "", "Resistencia (-)": "", "SI": "", "NO": "", "Observación": ""})
            df = add_row(df, {"Circuito": f"N°{i+1}", "Carga": f"{k+1}", "Configuración": "L-PE",
                              "Resistencia (+)": "", "Resistencia (-)": "", "SI": "", "NO": "", "Observación": ""})
            df = add_row(df, {"Circuito": f"N°{i+1}", "Carga": f"{k+1}", "Configuración": "L-N",
                              "Resistencia (+)": "", "Resistencia (-)": "", "SI": "", "NO": "", "Observación": ""})

        
        df = add_row(df, {"Circuito": "", "Carga": "", "Configuración": "",
                          "Resistencia (+)": "", "Resistencia (-)": "", "SI": "", "NO": "", "Observación": ""})

    df = add_row(df, {"Circuito": "", "Carga": "", "Configuración": "",
                      "Resistencia (+)": "", "Resistencia (-)": "", "SI": "", "NO": "", "Observación": ""})

    for _ in range(al):
        df = add_row(df, {"Circuito": "Alimentador", "Carga": "", "Configuración": "N-PE",
                          "Resistencia (+)": "", "Resistencia (-)": "", "SI": "", "NO": "", "Observación": ""})
        df = add_row(df, {"Circuito": "Alimentador", "Carga": "", "Configuración": "L-PE",
                          "Resistencia (+)": "", "Resistencia (-)": "", "SI": "", "NO": "", "Observación": ""})
        df = add_row(df, {"Circuito": "Alimentador", "Carga": "", "Configuración": "L-N",
                          "Resistencia (+)": "", "Resistencia (-)": "", "SI": "", "NO": "", "Observación": ""})

        df = add_row(df, {"Circuito": "", "Carga": "", "Configuración": "",
                          "Resistencia (+)": "", "Resistencia (-)": "", "SI": "", "NO": "", "Observación": ""})

    top = ["Circuito", "Circuito", "Medición", "Medición", "Medición", "CONFORME", "CONFORME", "CONFORME"]
    bottom = ["N° Circuito", "Carga", "Configuración", "Resistencia (+)", "Resistencia (-)", "SI", "NO", "Observación"]
    df.columns = pd.MultiIndex.from_arrays([top, bottom])


    out_dir = base_output_dir() / TEST_NAME
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "Tabla_continuidad.xlsx"

    df.index.name = ""
    df.to_excel(out_path, index=True)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    ws["A1"].value = None
    ws["A2"].value = None
    ws.column_dimensions["A"].hidden = True

    start_col = 2
    end_col = start_col + df.shape[1] - 1


    for c in range(start_col, end_col + 1):
        sub_cell = ws.cell(row=2, column=c)
        if isinstance(sub_cell, MergedCell):
            continue
        sub = sub_cell.value
        if sub is None or str(sub).strip() == "":
            sub_cell.value = None
            ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

    row_extra = df.columns.nlevels + 1
    ws.delete_rows(row_extra)



    wb.save(out_path)
    auto_ajustar_columnas(out_path)
    aplicar_bordes_excel(out_path)
    return out_path
