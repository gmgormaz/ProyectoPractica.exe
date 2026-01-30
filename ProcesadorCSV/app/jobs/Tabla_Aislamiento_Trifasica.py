from pathlib import Path
from itertools import combinations
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from app.output import base_output_dir
from app.output import aplicar_bordes_excel

TEST_NAME = "TT_Riso"


def _build_block_rows(circuito: str) -> list[dict]:

    fases = ["R", "S", "T", "N", "Ti"]
    pares = list(combinations(fases, 2))

    filas: list[dict] = []
    for a, b in pares:
        row = {"Circuito": circuito}  
        for f in fases:
            row[f] = ""
        row[a] = "X"
        row[b] = "X"
        row["Resistencia [MΩ]"] = ""
        row["SI"] = ""
        row["NO"] = ""
        row["N/A"] = ""
        row["Observación"] = ""
        filas.append(row)

    return filas


def run(n_circuitos: int, incluir_linea_general: bool = True) -> Path:
    if n_circuitos < 1:
        raise ValueError("n_circuitos debe ser >= 1.")

    fases = ["R", "S", "T", "N", "Ti"]
    block_len = len(list(combinations(fases, 2)))  # 10

    cols = [
        "Circuito", "R", "S", "T", "N", "Ti",
        "Resistencia [MΩ]", "SI", "NO", "N/A", "Observación"
    ]

    filas: list[dict] = []

    def add_separator():
        filas.append({c: "" for c in cols})

    if incluir_linea_general:
        filas += _build_block_rows("Línea General")
        add_separator()

    for i in range(1, n_circuitos + 1):
        filas += _build_block_rows(f"{i:02d}")
        add_separator()

    df = pd.DataFrame(filas, columns=cols)

    out_dir = base_output_dir() / TEST_NAME
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "TT_Riso.xlsx"

    df.to_excel(out_path, index=False)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center

    ws.freeze_panes = "A2"


    for r in range(2, ws.max_row + 1):
        for c in range(2, 7):
            ws.cell(r, c).alignment = center

    widths = {
        "A": 14,  
        "B": 5, "C": 5, "D": 5, "E": 5, "F": 5,  
        "G": 18,  
        "H": 6, "I": 6, "J": 6,  
        "K": 26, 
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


    row = 2

    def merge_circuit_cell(start_row: int, end_row: int):
        label = ws.cell(start_row, 1).value
        if label and str(label).strip():
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.cell(start_row, 1).alignment = center

    if incluir_linea_general:
        merge_circuit_cell(row, row + block_len - 1)
        row += block_len + 1  

    for _ in range(n_circuitos):
        merge_circuit_cell(row, row + block_len - 1)
        row += block_len + 1

    wb.save(out_path)

    aplicar_bordes_excel(out_path)

    return out_path
