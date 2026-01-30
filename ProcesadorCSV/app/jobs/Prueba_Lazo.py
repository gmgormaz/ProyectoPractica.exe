from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from app.output import make_output_path
from app.output import aplicar_bordes_excel
from app.output import auto_ajustar_columnas

TEST_NAME = "Prueba_Lazo"

ZS_MAX = {
    "B": {2:22, 4:11, 6:7.3, 10:4.4, 16:2.8, 20:2.2, 25:1.8, 32:1.4, 35:1.3, 40:1.1, 50:0.9, 63:0.7},
    "C": {2:11, 4:5.5, 6:3.65, 10:2.2, 16:1.4, 20:1.1, 25:0.9, 32:0.7, 35:0.65, 40:0.55, 50:0.45, 63:0.35},
    "D": {2:5.5, 4:2.8, 6:1.83, 10:1.1, 16:0.7, 20:0.55, 25:0.45, 32:0.34, 35:0.31, 40:0.27, 50:0.22, 63:0.17},
}

def run(input_path: Path, circuitos: list[dict]) -> Path:

    df = pd.read_csv(input_path)
    df = df[df["Test Function"].isin(["Prueba de lazo sin disparos"])].reset_index(drop=True)

    n = len(df)
    if len(circuitos) != n:
        raise ValueError("Cantidad de circuitos no coincide.")

    datos = {}
    for i in range(1, n + 1):
        In = int(circuitos[i-1]["In"])
        curva = circuitos[i-1]["curva"].upper()
        zsmax = ZS_MAX.get(curva, {}).get(In, np.nan)

        datos[i] = {"In": In, "curva": curva, "zsmax": zsmax}

    df["In[A]/Curva/KA"] = [f'{datos[i]["curva"]}/{datos[i]["In"]}' for i in range(1, n + 1)]
    df["Zs Max."] = [datos[i]["zsmax"] for i in range(1, n + 1)]

    out = df[[
        "In[A]/Curva/KA",
        "Primary Measurement",
        "Sub Measurement 2",
        "Zs Max.",
        "Sub Measurement 1",
        "Sub Measurement 3",
    ]].copy()

    pm = (
        out["Primary Measurement"]
        .astype(str)
        .str.replace(",", ".", regex=False)
        .str.extract(r"([-+]?\d*\.?\d+)")[0]
    )

    out["Zs [Ω]"] = pd.to_numeric(pm, errors="coerce")
    out["Zl [Ω]"] = out["Sub Measurement 2"]
    out["PEFC [A]"] = out["Sub Measurement 1"]
    out["PSC [A]"] = out["Sub Measurement 3"]
    out["IPCC [A]"] = (220 * 1.06) / out["Zs [Ω]"]

    out = out[[
        "In[A]/Curva/KA",
        "Zs [Ω]",
        "Zl [Ω]",
        "Zs Max.",
        "PEFC [A]",
        "PSC [A]",
        "IPCC [A]",
    ]]

    conforme = ["SI", "NO", "N/A", "Observación"]

    out_path = make_output_path(TEST_NAME, input_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    headers = ["Circuito"] + list(out.columns)

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c).value = h
        ws.cell(row=2, column=c).value = ""

    start = len(headers) + 1
    for i, sub in enumerate(conforme):
        ws.cell(row=1, column=start + i).value = "CONFORME"
        ws.cell(row=2, column=start + i).value = sub

    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start + len(conforme) - 1)

    bold = Font(bold=True)
    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=2, column=c).font = bold

    ws.freeze_panes = "A3"

    for r, row in enumerate(out.itertuples(index=False), start=3):
        ws.cell(row=r, column=1).value = r - 2
        for c, v in enumerate(row, start=2):
            ws.cell(row=r, column=c).value = v
    
    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    for r in (1, 2):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = bold
            cell.alignment = center

    wb.save(out_path)
    aplicar_bordes_excel(out_path)
    auto_ajustar_columnas(out_path)
    return out_path
