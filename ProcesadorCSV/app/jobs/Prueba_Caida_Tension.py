from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from app.output import make_output_path
from app.output import aplicar_bordes_excel
from app.output import auto_ajustar_columnas

TEST_NAME = "Prueba_Caida_Tension"

def run(input_path: Path, corrientes: list[int]) -> Path:
    df = pd.read_csv(input_path).reset_index(drop=True)

    df_lazo = df[df["Test Function"].isin(["Prueba de lazo sin disparos"])].copy().reset_index(drop=True)
    df_volt = df[df["Test Function"].isin(["Voltaje"])].copy().reset_index(drop=True)

    n = len(df_lazo)
    if len(df_volt) != n:
        raise ValueError(f"Cantidad no coincide: Lazo={n}, Voltaje={len(df_volt)}")
    if len(corrientes) != n:
        raise ValueError(f"Se esperaban {n} corrientes, llegaron {len(corrientes)}")

    out = pd.DataFrame()
    out["Circuito"] = list(range(1, n + 1))
    out["Voltaje"] = df_volt["Primary Measurement"].to_numpy()
    out["Z linea Ref. [Ω]"] = df_lazo["Primary Measurement"].to_numpy()
    out["Z linea [Ω]"] = ""              
    out["I nominal [A]"] = corrientes
    out["ΔV [%]"] = ""                   

    conforme = ["SI", "NO", "N/A", "Observación"]
    for s in conforme:
        out[s] = ""

    out_path = make_output_path(TEST_NAME, input_path, ext=".xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    cols_base = ["Circuito", "Voltaje", "Z linea Ref. [Ω]", "Z linea [Ω]", "I nominal [A]", "ΔV [%]"]

    for col_i, name in enumerate(cols_base, start=1):
        ws.cell(row=1, column=col_i).value = name
        ws.cell(row=2, column=col_i).value = ""

    start_conf = len(cols_base) + 1
    for j, sub in enumerate(conforme):
        ws.cell(row=1, column=start_conf + j).value = "CONFORME"
        ws.cell(row=2, column=start_conf + j).value = sub

    ws.merge_cells(
        start_row=1,
        start_column=start_conf,
        end_row=1,
        end_column=start_conf + len(conforme) - 1
    )

    bold = Font(bold=True)
    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c).font = bold
        ws.cell(row=2, column=c).font = bold

    ws.freeze_panes = "A3"

    for r, row in enumerate(out.itertuples(index=False), start=3):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c).value = val

    wb.save(out_path)
    aplicar_bordes_excel(out_path)
    auto_ajustar_columnas(out_path)
    
    return out_path
