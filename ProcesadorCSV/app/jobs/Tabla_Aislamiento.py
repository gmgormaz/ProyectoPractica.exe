from pathlib import Path
import openpyxl
from openpyxl.cell.cell import MergedCell
import pandas as pd
from app.output import base_output_dir
from app.output import aplicar_bordes_excel
from app.output import auto_ajustar_columnas

TEST_NAME = "Tabla_Aislamiento"


def add_row(df, row_dict):
    df.loc[len(df)] = row_dict
    return df


def nombre_salida_unico(path_out: Path) -> Path:
    if not path_out.exists():
        return path_out
    i = 1
    while True:
        candidato = path_out.with_stem(f"{path_out.stem}_{i}")
        if not candidato.exists():
            return candidato
        i += 1


def run(al: int, cir: int) -> Path:
    df = pd.DataFrame(columns=[
        "Circuito", "Configuración", "Resistencia", "SI", "NO", "Observación"
    ])

    for i in range(cir):
        df = add_row(df, {"Circuito": f"N°{i+1}", "Configuración": "N-PE", "Resistencia": "", "SI": "", "NO": "", "Observación": ""})
        df = add_row(df, {"Circuito": f"N°{i+1}", "Configuración": "L-PE", "Resistencia": "", "SI": "", "NO": "", "Observación": ""})
        df = add_row(df, {"Circuito": f"N°{i+1}", "Configuración": "L-N",  "Resistencia": "", "SI": "", "NO": "", "Observación": ""})
        df = add_row(df, {"Circuito": "", "Configuración": "", "Resistencia": "", "SI": "", "NO": "", "Observación": ""})

    df = add_row(df, {"Circuito": "", "Configuración": "", "Resistencia": "", "SI": "", "NO": "", "Observación": ""})

    
    for _ in range(al):
        df = add_row(df, {"Circuito": "", "Configuración": "N-PE", "Resistencia": "", "SI": "", "NO": "", "Observación": ""})
        df = add_row(df, {"Circuito": "", "Configuración": "L-PE", "Resistencia": "", "SI": "", "NO": "", "Observación": ""})
        df = add_row(df, {"Circuito": "", "Configuración": "L-N",  "Resistencia": "", "SI": "", "NO": "", "Observación": ""})
        df = add_row(df, {"Circuito": "", "Configuración": "", "Resistencia": "", "SI": "", "NO": "", "Observación": ""})

    top = ["Circuito", "Medición", "Medición", "CONFORME", "CONFORME", "CONFORME"]
    bottom = ["N° Circuito", "Configuración", "Resistencia", "SI", "NO", "Observación"]
    df.columns = pd.MultiIndex.from_arrays([top, bottom])

    out_dir = base_output_dir() / TEST_NAME
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = nombre_salida_unico(out_dir / "Tabla_aislamiento.xlsx")

    df.index.name = ""
    df.to_excel(out_path, index=True)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

 
    ws["A1"].value = None
    ws["A2"].value = None
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["A"].width = 0.1  


    start_col = 1
    end_col = df.shape[1]

    for c in range(start_col, end_col + 1):
        sub_cell = ws.cell(row=2, column=c)
        if isinstance(sub_cell, MergedCell):
            continue

        sub = sub_cell.value
        if sub is None or str(sub).strip() == "":
            sub_cell.value = None
            ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

    row_extra = df.columns.nlevels + 1
    ws.delete_rows(row_extra)

    wb.save(out_path)

    aplicar_bordes_excel(out_path)
    auto_ajustar_columnas(out_path)

    return out_path
