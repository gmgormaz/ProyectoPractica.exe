from pathlib import Path
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font,Alignment
from openpyxl.cell.cell import MergedCell
from app.output import make_output_path
from app.output import aplicar_bordes_excel
from app.output import auto_ajustar_columnas

TEST_NAME = "Prueba_Aislamiento"
umbral = 1000

def run(input_path: Path) -> Path:
    df = pd.read_csv(input_path)
    df = df.iloc[::-1].reset_index(drop=True)

    prueba2 = ['Prueba de aislamiento']
    df2 = df[df['Test Function'].isin(prueba2)].copy()

    if len(df2) == 0:
        raise ValueError("No hay filas con 'Prueba de aislamiento' en el CSV.")

    res = "Primary Measurement"
    no = "Notas"
    si = "Remark"

    num2 = (
        df2[res]
        .astype(str)
        .str.extract(r'([-+]?\d*\.?\d+)')[0]
        .astype(float)
    )

    df2[no] = np.where(num2 < 0.5, "X", "")
    df2[si] = np.where(num2 >= 0.5, "X", "")

    df2 = df2.rename(columns={
        "Primary Measurement": "Resistencia",
        "Level A": "Circuito",
        "Remark": " SI ",
        "Notas": " NO "
    })

    Ce = "Level B"
    alimentador = "Circuito"
    num = pd.to_numeric(df2[alimentador], errors="coerce")
    df2[alimentador] = np.where(num > umbral, "Al.", df2[Ce])

    desired_columns = ["Circuito", "Configuración", "Resistencia", " SI ", " NO "]
    df2 = df2[desired_columns].copy()

    top = ["Circuito","Configuraci+on","Resistencia","CONFORME","CONFORME"]
    bottom = [" "," "," ","SI","NO"]
    out = pd.DataFrame(df2.values, columns=pd.MultiIndex.from_arrays([top, bottom]))

    out_path = make_output_path(TEST_NAME, input_path, ext=".xlsx")

    out.index.name = ""
    out.to_excel(out_path, index=True)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    ws["A1"].value = None
    ws["A2"].value = None
    ws.column_dimensions["A"].hidden = True

    start_col = 2
    end_col = start_col + out.shape[1] - 1

    for c in range(start_col, end_col + 1):
        sub_cell = ws.cell(row=2, column=c)
        if isinstance(sub_cell, MergedCell):
            continue
        sub = sub_cell.value
        if sub is None or str(sub).strip() == "":
            sub_cell.value = None
            ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

    row_extra = out.columns.nlevels + 1
    ws.delete_rows(row_extra)
    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    for r in (1, 2):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = bold
            cell.alignment = center
    wb.save(out_path)
    auto_ajustar_columnas(out_path)
    aplicar_bordes_excel(out_path)
    return out_path
