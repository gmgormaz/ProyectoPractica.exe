
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment
from app.output import base_output_dir
from app.output import aplicar_bordes_excel
from app.output import auto_ajustar_columnas

TEST_NAME = "Tabla_Bucle_Falla"


def add_row(df: pd.DataFrame, row_dict: dict) -> pd.DataFrame:
    df.loc[len(df)] = row_dict
    return df


def _construir_circuitos(total_circuitos: int, monofasicos: set[int], incluir_linea_general: bool) -> list[str]:
    if total_circuitos < 1:
        raise ValueError("Circuitos totales debe ser >= 1.")
    if any(i < 1 or i > total_circuitos for i in monofasicos):
        raise ValueError("Hay circuitos monofásicos fuera de rango.")

    fases = ["R", "S", "T"]
    out = []

    if incluir_linea_general:
        out += [f"Línea General ({f})" for f in fases]

    for n in range(1, total_circuitos + 1):
        if n in monofasicos:
            out.append(f"{n}")  
        else:
            out += [f"{n:02d} ({f})" for f in fases]

    return out


def run(total_circuitos: int, monofasicos: list[int], incluir_linea_general: bool = True) -> Path:
    monofasicos_set = set(monofasicos)
    circuitos = _construir_circuitos(total_circuitos, monofasicos_set, incluir_linea_general)

    df = pd.DataFrame(columns=[
        "Circuito",
        "In[A]/Curva",
        "Zs [Ω]",
        "PEFC [A]",
        "PSC [A]",
        "IPCC [A]",
        "SI",
        "NO",
        "NA",
        "Observación",
    ])

    for c in circuitos:
        df = add_row(df, {
            "Circuito": c,
            "In[A]/Curva": "",
            "Zs [Ω]": "",
            "PEFC [A]": "",
            "PSC [A]": "",
            "IPCC [A]": "",
            "SI": "",
            "NO": "",
            "NA": "",
            "Observación": "",
        })

    top = [
        "Circuito",
        "Medición",
        "Medición",
        "Medición",
        "Medición",
        "Cálculo",
        "CONFORME",
        "CONFORME",
        "CONFORME",
        "CONFORME",
    ]
    bottom = [
        "",
        "In[A]/Curva",
        "Zs [Ω]",
        "PEFC [A]",
        "PSC [A]",
        "IPCC [A]",
        "SI",
        "NO",
        "NA",
        "Observación",
    ]
    df.columns = pd.MultiIndex.from_arrays([top, bottom])

    out_dir = base_output_dir() / TEST_NAME
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "Tabla_bucle_falla.xlsx"

    df.index.name = ""
    df.to_excel(out_path, index=True)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    ws["A1"].value = None
    ws["A2"].value = None
    ws.column_dimensions["A"].hidden = True

    start_col = 2
    end_col = start_col + df.shape[1] - 1

    for c in range(start_col, end_col + 1):
        sub_cell = ws.cell(row=2, column=c)
        if isinstance(sub_cell, MergedCell):
            continue

        sub = sub_cell.value
        if sub is None or str(sub).strip() == "":
            sub_cell.value = None
            ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for r in (1, 2):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = bold
            cell.alignment = center

    row_extra = df.columns.nlevels + 1
    ws.delete_rows(row_extra)

    wb.save(out_path)

    auto_ajustar_columnas(out_path)
    aplicar_bordes_excel(out_path)
    return out_path
